"""
Daily worklog -> docs/worklog.xlsx (gap: keep a running record of what we did).

Appends rows for entries that aren't already in the sheet (keyed by date+desc),
so it's safe to re-run any evening. New work each day: either add rows to
TODAY_ENTRIES below, or just let the git-commit rows auto-populate.

    python scripts/worklog.py
"""

import datetime
import subprocess
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "worklog.xlsx"
HEADERS = ["Date", "Time", "Category", "Module / Area", "What we did", "Commit / Status"]

# ---- manually-logged (non-commit) work. Add to this each day. ----
MANUAL = [
    # 2026-06-21 — gap brainstorm + survival/critical build pass (uncommitted)
    ("2026-06-21", "evening", "Docs", "Planning", "Gap analysis: vendor-side gaps, email/SMTP handling, future-updates/deploy model — written up", "docs/pending.md"),
    ("2026-06-21", "evening", "Docs", "Planning", "Client-vs-engineer brainstorm: 60 gaps across 7 rounds (critical/high/medium, daily-use, eng-catches, compliance, integrations, vendor ops)", "docs/pending.md"),
    ("2026-06-21", "evening", "Docs", "Planning", "UX/usability debt for HR daily use: U1-U12 (nav, approvals inbox, wizard, settings, action hierarchy, search, feedback...) parked as future", "docs/pending.md"),
    ("2026-06-21", "evening", "Fix", "Payroll", "Currency default INR: fixed $ fallback to Rupee in payroll/models/models.py (3 spots) for tenants without PayrollSettings", "payroll/models/models.py"),
    ("2026-06-21", "evening", "Feature", "Ops/Infra", "#8 Nightly DB backup script: pg_dump + media tar, 14-day rotation, reads .env DB creds", "scripts/backup_db.sh"),
    ("2026-06-21", "evening", "Feature", "Ops/Infra", "#12 One-command deploy script: safety-backup -> git pull -> migrate -> collectstatic -> restart services", "scripts/deploy.sh"),
    ("2026-06-21", "evening", "Feature", "Billing", "#5 Razorpay webhook: verify_webhook (HMAC of raw body) + razorpay_webhook view + pay/webhook/ route; activates plan server-side from receipt", "subscriptions/billing.py, views.py, client_urls.py"),
    ("2026-06-21", "evening", "Feature", "Subscriptions", "#7 Trial-ending in-app banner (<=7d, red <=2d) in index.html + trial_days_left context", "skylinx_theme/templates/index.html, subscriptions/context_processors.py"),
    ("2026-06-21", "evening", "Feature", "Subscriptions", "#7 notify_trial_ending command: emails admins at 7/3/1 days before trial end", "subscriptions/management/commands/notify_trial_ending.py"),
    ("2026-06-21", "evening", "Feature", "Subscriptions", "#30 Multi-admin / account recovery: company_admins view + admins.html + route + profile link; promote/revoke admin, blocks removing last admin", "subscriptions/views.py, templates/subscriptions/admins.html"),
    ("2026-06-21", "evening", "Feature", "Ops/Infra", "Daily worklog generator -> docs/worklog.xlsx", "scripts/worklog.py"),
]


def git_rows():
    """One row per commit from 2026-06-19 onward."""
    out = subprocess.run(
        ["git", "log", "--since=2026-06-19 00:00",
         "--date=format:%Y-%m-%d|%H:%M", "--pretty=format:%ad|%h|%s"],
        capture_output=True, text=True, cwd=OUT.parent.parent,
    ).stdout.strip().splitlines()
    rows = []
    for line in out:
        parts = line.split("|", 3)
        if len(parts) < 4:
            continue
        date, time, sha, subj = parts
        rows.append((date, time, "Commit", "git", subj, sha))
    return rows


def load_or_new():
    if OUT.exists():
        wb = openpyxl.load_workbook(OUT)
        return wb, wb.active
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Worklog"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="2563EB")
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    return wb, ws


def main():
    wb, ws = load_or_new()
    # existing keys (date + description) to avoid duplicates on re-run
    seen = {(r[0].value, r[4].value) for r in ws.iter_rows(min_row=2)}
    rows = MANUAL + git_rows()
    rows.sort(key=lambda r: (r[0], r[1]), reverse=True)  # newest first
    added = 0
    for r in rows:
        if (r[0], r[4]) in seen:
            continue
        ws.append(r)
        seen.add((r[0], r[4]))
        added += 1
    # widths
    widths = [12, 9, 11, 22, 90, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
    OUT.parent.mkdir(exist_ok=True)
    wb.save(OUT)
    print(f"worklog.xlsx: +{added} rows, {ws.max_row - 1} total -> {OUT}")


if __name__ == "__main__":
    main()
